#!/bin/env python
# /c/AppData/Code/venv/py/dc/dc_sup.py
import os
import json
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.cell.text import InlineFont, Text
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.cell.rich_text import TextBlock, CellRichText
# from typing import Callable
# from openpyxl.styles.alignment import Alignment
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Font, PatternFill, Border, Side
class DC_SUP:
    verbose=0
    excel_max_width=180
    capa_tables={
        'CONTROLE DE VERSÃO':{
            'Versão':'Alteração',
            '1.0':'Usada intra equipes para configuração maunal',
            '2.0':'DC gerada automaticamente.\nOBS 1: ITX, atentar para conbine-se intra equipes de que uma DC só poderá ter 1 Servico, com 1 Abrangência e com 1 CN exceto quando Estadual/Nacional.\nOBS 2: Eventuais colunas após Município_a e antes de Obs podem ser adicionadas para atender outras especificações.',
            '3.0':'Alteração interna. funções em Banco e encaminhamento NGN',
            '4.0':'Automática via eVoice',
        },
        'Discrminação: Tabela Encaminhamento':{
            'Campo':'Descrição',
            'Origem':'Define a origem da ligação\nSMP: (Móvel) Serviço Móvel Pessoal\nSTFC: (Fixa) Serviço Telefônico Fixo Comutado\nVoLTE: (Voz 4G)Voz sobre LTE\nCORP: (Corporativo) Rede Classe 5\nLIVE: (Live Tim) Telefonia Fixa sobre Fibra\n[vazio]: Sem Corbetura ou erro',
            'Abrangência':'Define a abrangência da tradução.\nNacional: Todo Brasil. *Não gera tabela de Células\nEstadual: A todo a UF 1*Todas os CNs serão mostrados, 2*Não gera tabela de Células\nANF: Por CN\nÁrea Local: Por Área Local\nMunicipal: Por Cidade (Município)\nLocalidade: [Depreciado] Por Localidade (Parte de um Município)\nEmergency Center: [Depreciado] Por área de Emergência (Parte de uma localidade/Bairros)',
            'Serviço':'Servico Tridígito a ser traduzido',
            'Tipo Serv':'Tipo do Serviço\nSUP: Serviços de Utilidade Pública\nSPE: Serviços Públicos Emergenciais\nSAS: Serviços de Apoio ao STFC\nSTF: (103*) Serviços de Telefonia Fixa\nSTM: (105*) Serviços de Telefonia Móvel\nSTA: (106*) Serviços de Televisão por Assinatura',
            'Tradução':'Para qual número a operadora destino (OLO) pediu para traduzir o número. *Todos os elementos do número traduzido são separados por espaço podendo ser comparados com o campo Formato OLO',
            'Formato OLO':'Formato do númeor de tradução como OLO definiu separando todos os elementos por espaço. *O formato pode ter várias combinações com os elementos abaixo ex: 0 CN N8, 0 CN SE, 0 CN SE CG, 0 CG SE etc\nCN: (CNb/CNd) Código Nacional/ANF de destino, 2 dígitos de 1 a 9. É o CN junto ao número de lista\nCNb: idem CN\nN8: 8 Digitos, um número de lista fixo\nN9: 9 Digitos, um número de lista móvel\nCNG: Código Não Geográfico, um 0800 por exemplo\nSE: Serviço. *Em casos especiais, como 112 e 911, pode ser convertido para outro ex: 190\nCG: Cifra Guia, um número de 1 a 6 dígitos para designar na OLO para qual região aternder\nSCM: Short Code Massivo (pode ser tratado como um CG apenas), um formato especial interno, ex: 017003001\nCNL: Código Nacional de Localidade. É uma CG com 5 dígitos que designa uma localidade. *Pode ser tratado naturalmente como CG\n[digitos]: Dígitos de 0 a 9\n[letras]: Letras de A, B, C, D, e E que são traduzidas respectivamente para #10, #11, #12, #13, #14',
            'Tipo TR':'Tipo da Tradução\nN8: Número de Lista\nSE: Serviço + Cifra Guia\nCNG: Código Não Geográfico',
            'Tarifação':'Onde será tarifada a ligação\nNP: Nimguém Paga\nAP: A Paga\nBP: B Paga',
            'RN2':'RN2 AXRN onde A0=Não Portado, A1=Portado',
            'ROP':'ROP da ligação\nROPa: ROP de origem quando Tipo TR=SE\nROPb: ROP quando Tipo TR=N8\nROPd: ROP dummy quando Tipo TR=CNG. *00000',
            'Central Origem':'Central que faz parte de um passo da chamada que trata o número e passa adiante\nZ[CNL*]: Centrais Ericsson\nVSC[*]: Centrais NGN\n[outras]: Centrais Huawei IMS/CL5',
            'Central Destino':'Central que faz parte de um passo da chamada que recebe o número da Central de Origem ou finaliza entrega quando OLO. *Se houver mais de uma Central de Destino essas serão separadas por ","\n[Idem Origem]: Idem Central Origem\nOLO: Entrega chamada para Operadora de Destino',
            'Rota Destino':'Rota(s) utilizada(s) para fazer a entrega de Origem para Destino\n[Crítica 1]: Se houver mais de uma rota esta será separada por "/" formando um grupo de rotas\n[Crítica 2]: Se houver mais de uma Central de Destino os grupos de rotas serão separados por ","',
            'Formato Envio':'Formato de envio da chamada pela rota intra centrais deve seguir um padrão estipulado por ITX em acordo com Configuração, estes cadastrados no ARQUIVÃO e ROBOC\n[Idem Formato OLO]: Idem Formato OLO\n[()]: Parenteses formam grupos que são adicionados ou não dependendo da regra de negócio. *Os grupos podem ser reagrupados, ou seja, parentes dentro de parenteses\nX: Qualquer dígito de tamanho 1. Seria o mesmo que N1\nRN: É o código da OLO de 3 dígitos\nRN2: Formato AX+RN onde AX determina A0 para número não portado e A1 para portado\nA8RN: O mesmo que A8 RN. Utilizado para transbordo.\nN8/N9: Duas possibilidades: ou N8 ou N9\nCNa: (CNo) CN de origem que é de onde foi discado o número\nROPa: ROPa Origem\nROPb: ROPb Destino\nROPd: ROPb Dummy utilizado para CNG\nCSP: CSP 041\n,: Separação para escolhas de formatos diferentes\nMSRN: Mobile Subscriber Roaming Number\n[etc]: Outros elementos de formatos não mapeados',
            'CNa':'Código Nacional de Origem',
            'Cod_CNL_a':'Código Nacional de Localidade de Origem.\n[Crítica 1]: Utilizado para vincular a tabela de SERVIÇOS com CELLS CS\n[Crítica 2]: Cod_CNL, acima de abrangências Municipais é representada pela Localidade mais importante.\n[Crítica 3]: Cod_CNL pode não corresponder a Sigla_CNL devido a Crítica acima',
            'CNL_a':'Corresponde a Sigla da Localidade de Origem requerida pela OLO para entregar a tradução',
            'AL_a':'Área Local de CNL_a',
            'UF_a':'Unidade Federativa de CNL_a',
            'Município_a':'Município de CNL_a',
            'Obs':'Campo destinado a mais informações',
        },
        'Discrminação: Tabela Cells':{
            'Campo':'Descrição',
            'Cell':'Id da Célula da ERB',
            'CGI':'Common Gateway Interface/Interface Comum de Porta de entrada (Id da ERB)',
            'Tecnologia':'Tecnologia da Célula\n2G: 2G\n3G: 3G\n4G: 4G',
            'Cod_CNL':'Cod_CNL que vincula as tabelas de SERVIÇOS e CELLS CS',
            'Sigla_CNL':'Sigla CNL correspondente a célula',
        },
    }
    def __init__(self):
        self.level=0
        self.path=os.path.dirname(os.path.abspath(__file__))
        self.sheets={
            'capa':{
                'title':'CAPA',
                'fn_format':self.excel_format_capa,
                },
            'encam':{
                'title':'SERVICOS',
                'fn_format':self.excel_insert_table,
                },
            'cgi':{
                'title':'CELLS CS',
                'fn_format':self.excel_insert_table,
                },
        }
        self.capa={}
        self.dc_dict:dict[str,dict] = self.get_json_file(f"{self.path}/data/dict-1.0.0.json")
        self.encam_dict:dict[str,dict[str,dict[str,dict[str,dict[str,dict]]]]]=self.get_json_file(f"{self.path}/data/encam.json")
        self.df:dict={} # Reg,[header,encam,cgi],pd.DataFrame
        self.data_extract_encam()
        self.data_extract_cgi(f"{self.path}/data/cgi.json")
        self.data_extract_capa()
        
        for reg in self.df:
            self.wb = self.excel_create_workbook()
            for sheet in self.df[reg]:
                df_item:pd.DataFrame=self.df[reg][sheet]
                if not len(df_item):
                    self.show(f"{reg}:{sheet}: Nenhum dado encontrado")
                    continue

                title=self.sheets[sheet]['title']
                fn_format=self.sheets[sheet]['fn_format']
                ws=self.convert_to_excel(sheet,df_item)
                fn_format(title,ws)
                # self.show_done(title,df_item)

            self.excel_save(f"{self.path}/xlsx/dc_{reg}.xlsx")

    def data_extract_capa(self):
        data={}
        for idSup,sup_data in self.encam_dict['data']['sup'].items():
            servico=sup_data['Servico']
            servico_descr=sup_data['Servico_Descricao']
            idTipoSrv=str(sup_data['idTipoSrv'])
            servico_tipo=self.dc_dict['tipo_serv'][idTipoSrv]['TipoSrv']
            idAbrangencia=str(sup_data['idAbrangencia'])
            abrangencia=self.dc_dict['abrangencia'][idAbrangencia]['Abrangencia']
            rn1=sup_data['RN1']
            olo=self.encam_dict['data']['rn1'][f'{rn1}']['Prestadora']
            # olo=self.dc_dict['rn1'][f'{rn1}']['OLO']
            # rn1_grp=self.dc_dict['rn1'][f'{rn1}']['RN1_Grp']
            # cnpj=self.dc_dict['rn1'][f'{rn1}']['CNPJ']
            
            action_key=f'Abertura de {servico_tipo} {servico}: {olo} ({rn1})'
            descr_key=f'{servico_tipo} {servico}: {abrangencia}' # (CNLs)
            for cod_cnl,tipo_orig_data in sup_data['cnl'].items():
                cnl_data=self.encam_dict['data']['cnl'][f'{cod_cnl}']
                sigla_cnl=cnl_data['Sigla_CNL']
                # municipio=cnl_data['Municipio']
                idAL=str(cnl_data['idAL'])
                al_data=self.encam_dict['data']['al'][idAL]
                reg=al_data['Reg']
                uf=al_data['UF']
                cn=str(al_data['CN'])
                if not data.get(reg):
                    data[reg]={
                        'acao':{}, # Abertura de SE 190:OI S.A. - EM RECUPERACAO JUDICIAL
                        'descr':{}, # SE 190: Municipal(CEO,CPG,CRC,CVY,CCP,PGC)
                    }
                if not data[reg]['descr'].get(descr_key):
                    data[reg]['descr'][descr_key]={}
                data[reg]['acao'][action_key]=action_key
                if idAbrangencia=="3": # Estadual
                    data[reg]['descr'][descr_key][uf]=uf
                elif idAbrangencia=="6": # CN
                    data[reg]['descr'][descr_key][cn]=cn
                elif idAbrangencia!="4": # Nacional
                    data[reg]['descr'][descr_key][sigla_cnl]=sigla_cnl
        
        for reg in data:
            for descr_key,descr_data in data[reg]['descr'].items():
                data[reg]['descr'][descr_key]=f'{descr_key} ({",".join(descr_data.values())})'
            sx_a=''
            tg_a=''
            if self.capa.get(reg):
                # SMP:(ZBHE04,ZBHE05,ZBSA05,ZCEM02,ZCEM03,ZRJO03,VSC1N,VSC2N)
                # STFC:(ZBHE04,ZBHE05,ZBSA05,ZCEM02,ZCEM03,ZRJO03)
                sx_a=[]
                tg_a=[]
                for tipoOrig,to_data in self.capa[reg].items():
                    if len(to_data['sx_a']):
                        sx_a.append(f"{tipoOrig}: ({",".join(to_data['sx_a'].values())})")
                    if len(to_data['tg_a']):
                        tg_a.append(f"{tipoOrig}: ({",".join(to_data['tg_a'].values())})")
                sx_a='\n'.join(sx_a)
                tg_a='\n'.join(tg_a)
            capa=self.build_capa_data(
                dc=self.encam_dict['header']['dc'],
                owner=self.encam_dict['header']['Colaborador'],
                dt=self.encam_dict['header']['dt_ger'],
                tg_a=tg_a,
                sx_a=sx_a,
                action='\n'.join(data[reg]['acao'].values()),
                descr='\n'.join(data[reg]['descr'].values()),
            )
            self.df[reg]['capa']=pd.DataFrame(capa)
    
    def build_capa_data(self,dc:str='',owner:str='',dt:str='',tg_a:str='',sx_a:str='',tg_b:str='',sx_b:str='',action:str='',descr:str=''):
        capa=[]
        # Colunas A, B, C, D, E
        capa.append(['DOCUMENTO DE CONFIGURAÇÃO','','','',''])
        capa.append(['','','','',''])
        
        capa.append(['DC:',dc,'','RESPONSÁVEL:',owner])
        capa.append(['','','','DATA:',dt])
        capa.append(['LADO A','ROTA:',tg_a,'CENTRAL:',sx_a])
        capa.append(['LADO B','ROTA:',tg_b,'CENTRAL:',sx_b])
        capa.append(['','','','',''])
        
        capa.append(['AÇÃO:',action,'','',''])
        capa.append(['DESCRIÇÃO:',descr,'','',''])
        
        for sub_title,tbl in self.capa_tables.items():
            capa.append(['','','','',''])
            capa.append([sub_title,'','','',''])
            for k,v in tbl.items():
                capa.append([k,v,'','',''])
        return capa
    
    def data_extract_encam(self):
        # dict[str,dict]
        # Origem,Abrangencia,Servico,Tipo Serv,Traducao,Formato OLO,Tipo TR,Tarifacao,RN2,ROP,AL,Central Origem,Central Destino,Rota Destino,Formato Envio,CN_a,Cod_CNL_a,CNL_a,AL_a,UF_a,Municipio_a,Obs
        if not self.encam_dict: return
        
        data:dict[str,list] = {}
        arr_error:dict=self.dc_dict['error']
        
        self.level_open(f'Encam')
        for sx_a, sx_b_data in self.encam_dict['forwarding'].items():
            self.level_open(f'sx_a: {sx_a}')
            for sx_b, rn1_data in sx_b_data.items():
                self.level_open(f'sx_b: {sx_b}')
                for rn1, rota_summary_data in rn1_data.items():
                    self.level_open(f'rn1: {rn1}')
                    for rota_summary, formato_summary_data in rota_summary_data.items():
                        self.level_open(f'rota_summary: {rota_summary}')
                        for formato_summary, traducao_data in formato_summary_data.items():
                            self.level_open(f'formato_summary: {formato_summary}')
                            for traducao, encam_data in traducao_data.items():
                                self.level_open(f'traducao: {traducao} (ord: {encam_data['ord']})')

                                if encam_data['ord']==0: tipo_ord='Acesso'
                                elif encam_data['ord']==-1: tipo_ord='OLO'
                                else:
                                    tipo_ord='Rotas Internas'
                                    continue
                                
                                arr_obs=[]
                                error=encam_data['error']
                                mark,error_str='',''
                                if error: 
                                    mark='*'
                                    error_str=f"ERROR: {error}-{arr_error[f'{error}']['desc']}"
                                    # arr_obs.append(error_str)
                                    self.evel_item(error_str)
                                    
                                sup_data:dict=encam_data['sup']
                                orig_data:dict=encam_data['orig']
                                # trunk_type_data:dict=encam_data['trunk_type']
                                rotas_data:dict[str,dict]=encam_data['rotas']
                                local_data:dict[str,dict]=encam_data['local']
                                arr_al_b:dict[str,dict]=encam_data['al_b']
                                
                                arr_rota_capa={}
                                if rotas_data:
                                    for id_encam_type, rotas_data_det in rotas_data.items():
                                        self.level_open(f'id_encam_type: {id_encam_type}')
                                        arr_rota,arr_carga,arr_formato=[],[],[]
                                        for rota, arr_rota_det in rotas_data_det.items():
                                            self.level_open(f'rota: {rota}')
                                            if not rota: continue
                                            carga=arr_rota_det[0]
                                            formato=arr_rota_det[1]
                                            if id_encam_type=='1':
                                                arr_rota_capa[rota]=rota
                                            arr_rota.append(rota)
                                            arr_carga.append(f'{carga}' if carga else '-')
                                            arr_formato.append(f'{formato}' if formato else '-')
                                            self.level_close('rota')
                                        self.level_close('id_encam_type')
                                        
                                        arr_encam_type:dict=self.dc_dict['encam_type'][id_encam_type]
                                        encam_type:str=arr_encam_type['Quando']+(' Transbordo' if arr_encam_type['Transbordo'] else ' Normal')
                                        # print([arr_rota,arr_carga,arr_formato])
                                        arr_obs.append(f'- {encam_type}: {'/'.join(arr_rota)}({'/'.join(arr_carga)})[{';'.join(arr_formato)}]')
                                obs:str=' \n'.join(arr_obs)
                                    
                                for idSup, servico in sup_data.items():
                                    self.level_open(f'idSup[{idSup}]: servico {servico}')
                                    sup:dict=self.encam_dict['data']['sup'][idSup]

                                    idAbrangencia=str(sup['idAbrangencia'])
                                    idTipoSrv=str(sup['idTipoSrv'])
                                    abrangencia:str=self.dc_dict['abrangencia'][idAbrangencia]['Abrangencia']
                                    tipo_serv:str=self.dc_dict['tipo_serv'][idTipoSrv]['TipoSrv']
                                    
                                    arr_sup_cnl:dict=sup['cnl']
                                    
                                    self.evel_item(f'idAbrangencia: {idAbrangencia}-{abrangencia}')
                                    self.evel_item(f'idTipoSrv....: {idTipoSrv}-{tipo_serv}')
                                    
                                    # Localidades Ponta A
                                    for cod_cnl_a,idTipoOrig_data in arr_sup_cnl.items():
                                        self.level_open(f'Cod_CNL_a[{cod_cnl_a}]: {cod_cnl_a}')
                                        cnl_data=self.encam_dict['data']['cnl'][cod_cnl_a]
                                        sigla_cnl_a=cnl_data['Sigla_CNL']
                                        municipio_a=cnl_data['Municipio']
                                        idAL_a=str(cnl_data['idAL'])
                                        al_data=self.encam_dict['data']['al'][idAL_a]
                                        al_a=al_data['AL']
                                        rop_a=al_data['ROP']
                                        cn_a=al_data['CN']
                                        uf_a=al_data['UF']
                                        reg_a=al_data['Reg']
                                        

                                        for idTipoOrig,arr_idTipoOrig in idTipoOrig_data.items():
                                            if idTipoOrig in orig_data:
                                                cn_b, idAL_b, rn1_b, idPortado, grp, _traducao = arr_idTipoOrig
                                                reg_b, uf_b, cn_b, al_b, rop_b =arr_al_b[f'{idAL_b}']
                                                arr_tipoOrig:dict=self.dc_dict['tipo_orig'][idTipoOrig]
                                                tipoOrig=arr_tipoOrig['TipoOrig']
                                                self.level_open(f'tipoOrig[{idTipoOrig}]: {tipoOrig}')
                                                
                                                TrType:str=sup[arr_tipoOrig['TrType']]
                                                idTarifacao=str(sup[arr_tipoOrig['idTarifacao']])
                                                tarifacao=self.dc_dict['tarifacao'][idTarifacao]['Tarif']
                                                tipo_tr:str=self.dc_dict['tr_types'][TrType]
                                                self.evel_item(f'idTarifacao: {idTarifacao}-{tarifacao}')
                                                self.evel_item(f'TrType.....: {TrType}-{tipo_tr}')
                                                
                                                data_item={
                                                    'Origem': mark+tipoOrig,
                                                    'Abrangencia': abrangencia, # arr_abrangencia
                                                    'Servico': servico,
                                                    'Tipo Serv': tipo_serv,
                                                    'Traducao': traducao,
                                                    'Formato OLO': TrType,
                                                    'Tipo TR': tipo_tr,
                                                    'Tarifacao': tarifacao,
                                                    'RN1': f'{rn1_b}',
                                                    'ROP': f'{rop_b}',
                                                    'AL': al_b,
                                                    'Central Origem': sx_a,
                                                    'Central Destino': sx_b,
                                                    'Rota Destino': rota_summary,
                                                    'Formato Envio': formato_summary,
                                                    'CN_a': f'{cn_a}',
                                                    'Cod_CNL_a': f'{cod_cnl_a}',
                                                    'CNL_a': sigla_cnl_a,
                                                    'AL_a': al_a,
                                                    'UF_a': uf_a,
                                                    'Municipio_a': municipio_a,
                                                    'Obs': obs,
                                                    'Ord': f'# {encam_data['ord']} {tipo_ord}',
                                                    'Error': error_str,
                                                }
                                                reg=reg_b # Regional de A ou de B?
                                                if not self.capa.get(reg):
                                                    self.capa[reg]={}
                                                if not self.capa[reg].get(tipoOrig):
                                                    self.capa[reg][tipoOrig]={
                                                        'sx_a':{},
                                                        'tg_a':{},
                                                    }
                                                if not data.get(reg): data[reg]=[]
                                                
                                                data[reg].append(data_item)
                                                self.capa[reg][tipoOrig]['sx_a'][sx_a]=sx_a
                                                self.capa[reg][tipoOrig]['tg_a']|=arr_rota_capa
                                                self.level_close('tipoOrig')
                                        self.level_close('Cod_CNL_a')
                                    
                                    # Outra Abordagem Depreciada
                                    # for idTipoOrig, tipoOrig in orig_data.items():
                                        # self.level_open(f'tipoOrig[{idTipoOrig}]: {tipoOrig}')
                                        # arr_tipoOrig:dict=self.dc_dict['tipo_orig'][idTipoOrig]
                                        # TrType:str=sup[arr_tipoOrig['TrType']]
                                        # idTarifacao=str(sup[arr_tipoOrig['idTarifacao']])
                                        # tarifacao=self.dc_dict['tarifacao'][idTarifacao]['Tarif']
                                        # tipo_tr:str=self.dc_dict['tr_types'][TrType]
                                        # self.evel_item(f'idTarifacao: {idTarifacao}-{tarifacao}')
                                        # self.evel_item(f'TrType.....: {TrType}-{tipo_tr}')
                                        
                                        
                                        # for reg, uf_data in local_data.items():
                                            # self.level_open(f'Reg: {reg}')
                                            # if not self.capa.get(reg):
                                            #     self.capa[reg]={}
                                            # if not self.capa[reg].get(tipoOrig):
                                            #     self.capa[reg][tipoOrig]={
                                            #         'sx_a':{},
                                            #         'tg_a':{},
                                            #     }
                                            
                                            # self.capa[reg][tipoOrig]['sx_a'][sx_a]=sx_a
                                            # self.capa[reg][tipoOrig]['tg_a']|=arr_rota_capa
                                            # arr_uf,arr_cn,arr_al,arr_rop,arr_cnl,arr_cod_cnl,arr_ct={},{},{},{},{},{},{}
                                            # for uf, cn_data in uf_data.items():
                                            #     self.level_open(f'UF: {uf}')
                                            #     arr_uf[uf]=uf
                                            #     for cn, al_data in cn_data.items():
                                            #         self.level_open(f'CN: {cn}')
                                            #         arr_cn[cn]=f'{cn}'
                                            #         for idAL, idAL_data in al_data.items():
                                            #             self.level_open(f'idAL: {idAL}-{idAL_data["AL"]}')
                                            #             arr_al[idAL]=idAL_data["AL"]
                                            #             arr_rop[idAL]=f'{idAL_data["ROP"]}'
                                            #             for cod_cnl, Sigla_CNL in idAL_data['CNL'].items():
                                            #                 self.level_open(f'Cod_CNL: {cod_cnl}')
                                            #                 arr_cod_cnl[cod_cnl]=f'{cod_cnl}'
                                            #                 arr_cnl[cod_cnl]=Sigla_CNL
                                            #                 arr_ct[cod_cnl]=f'{self.encam_dict["data"]["cnl"][cod_cnl]["Municipio"]}'
                                            #                 self.level_close('Cod_CNL')
                                            #             self.level_close('idAL')
                                            #         self.level_close('CN')
                                            #     self.level_close('UF')
                                            # data_item={
                                            #     'Origem': mark+tipoOrig,
                                            #     'Abrangencia': abrangencia, # arr_abrangencia
                                            #     'Servico': servico,
                                            #     'Tipo Serv': tipo_serv,
                                            #     'Traducao': traducao,
                                            #     'Formato OLO': TrType,
                                            #     'Tipo TR': tipo_tr,
                                            #     'Tarifacao': tarifacao,
                                            #     'RN1': rn1,
                                            #     'ROP': ','.join(arr_rop.values()),
                                            #     'AL': ','.join(arr_al.values()),
                                            #     'Central Origem': sx_a,
                                            #     'Central Destino': sx_b,
                                            #     'Rota Destino': rota_summary,
                                            #     'Formato Envio': formato_summary,
                                            #     'CN_a': ','.join(arr_cn.values()),
                                            #     'Cod_CNL_a': ','.join(arr_cod_cnl.values()),
                                            #     'CNL_a': ','.join(arr_cnl.values()),
                                            #     'AL_a': ','.join(arr_al.values()), # avaliar se é A ou B
                                            #     'UF_a': ','.join(arr_uf.values()),
                                            #     'Municipio_a': ','.join(arr_ct.values()),
                                            #     'Obs': obs,
                                            #     'Ord': f'# {encam_data['ord']} {tipo_ord}',
                                            #     'Error': error_str,
                                            # }
                                            # # ,,,,,
                                            # if not data.get(reg): data[reg]=[]
                                            # data[reg].append(data_item)
                                            # self.level_close('Reg')
                                        # self.level_close('tipoOrig')
                                    self.level_close('idSup')
                                self.level_close('traducao')
                            self.level_close('formato_summary')
                        self.level_close('rota_summary')
                    self.level_close('rn1')
                self.level_close('sx_b')
            self.level_close('sx_a')
        self.level_close('Encam')
        
        for reg in data:
            if not self.df.get(reg):
                self.df[reg]={}
            self.df[reg]['capa']={}
            data[reg]=pd.DataFrame(data[reg])
            # self.show_done(f'{reg} Encam',data[reg])
            data[reg]=self.group_encam_data(data[reg])
            data[reg]=data[reg].sort_values(by=['Origem', 'Central Destino'], ascending=[True, True])

            self.show_done(f'{reg} Encam',data[reg])
            self.df[reg]['encam']=data[reg]

    def group_encam_data(self,df:pd.DataFrame):
        all_cols = df.columns.tolist()
        # Criar uma nova coluna para indicar se existe erro
        df['Tem_Error'] = df['Error'].notna() & (df['Error'].str.strip() != '')
        
        # Definir as colunas-chave para agrupamento
        group_cols = [
            'Servico',
            'Traducao',
            'RN1',
            'Central Origem',
            'Tem_Error'
        ]
        
        # Colunas para aplicar distinct (excluindo as colunas-chave)
        colunas_distinct = [col for col in all_cols if col not in group_cols]
        
        def aplicar_distinct(series):
            """Função para obter valores distintos de uma série, removendo valores nulos"""
            valores_unicos = series.dropna().unique()
            # Remove strings vazias
            valores_unicos = [v for v in valores_unicos if str(v).strip() != '']
            return ';\n'.join(list(valores_unicos)) if len(valores_unicos) > 0 else ''
        
        # Criar dicionário de agregação
        agg_dict = {}
        
        # Para colunas-chave, usar 'first' (já que serão únicas no grupo)
        for col in group_cols:
            if col != 'Tem_Error' and col in df.columns:
                agg_dict[col] = 'first'
        
        # Para demais colunas, aplicar distinct
        for col in colunas_distinct:
            if col in df.columns:
                agg_dict[col] = aplicar_distinct
        
        # Realizar o agrupamento
        resultado = df.groupby(group_cols, as_index=False).agg(agg_dict)
        
        # Remover a coluna Tem_Error do resultado final (era apenas para agrupamento)
        if 'Tem_Error' in resultado.columns:
            resultado = resultado.drop('Tem_Error', axis=1)

        # Converter listas em strings, dá problema para converter para excel se não fizer
        # for col in resultado.columns:
        #     if resultado[col].dtype == 'object':
        #         resultado[col] = resultado[col].apply(
        #             lambda x: '; '.join(map(str, x)) if isinstance(x, list) else str(x)
        #         )
        
        def concatenar_colunas_se_nao_vazio(row):
            valores = []
            if pd.notna(row['Ord']) and str(row['Ord']).strip():
                valores.append(str(row['Ord']).strip())
            if pd.notna(row['Error']) and str(row['Error']).strip():
                valores.append(str(row['Error']).strip())
            if pd.notna(row['Obs']) and str(row['Obs']).strip():
                valores.append(str(row['Obs']).strip())
            return '\n'.join(valores)

        # concatenar Ord+Error+Obs=Obs
        # resultado['Obs'] = resultado.apply(concatenar_colunas_se_nao_vazio, axis=1)
        # resultado = resultado[ [
        #     'Origem', 'Abrangencia', 'Servico', 'Tipo Serv', 'Traducao', 
        #     'Formato OLO', 'Tipo TR', 'Tarifacao', 'RN1', 'ROP', 'AL', 
        #     'Central Origem', 'Central Destino', 'Rota Destino', 'Formato Envio',
        #     'CN_a', 'Cod_CNL_a', 'CNL_a', 'AL_a', 'UF_a', 'Municipio_a', 
        #     'Obs'
        # ] ]
        resultado = resultado[all_cols]
        
        # resultado = resultado.drop('Ord', axis=1)
        # resultado = resultado.drop('Error', axis=1)
        
        return resultado
        return df

    def data_extract_cgi(self,file:str):
        """
        Extrai dados do JSON CGI em formato tabular
        
        - from: Sigla_CNL,Cod_CNL,ERN,G,EA,EndId,SiteId,CGI,Celula,EC,ERIND,idDevice=Device
        - to: CGI,RAT=G,Sigla_CNL,EC,ERN,Devices order(RAT, Sigla_CNL) 
        """
        
        json_dict:dict[str,dict[str,dict[str,dict[str,dict[str,dict[str,dict[str,dict[str,dict[str,dict[str,dict[str,dict[str,dict]]]]]]]]]]]]=self.get_json_file(file)
        if not json_dict: return
        
        self.level_open(f'CGI')
        for reg, sigla_cnl_data in json_dict.items():
            data = []
            self.level_open(f'reg: {reg}')
            for sigla_cnl, cnl_data in sigla_cnl_data.items():
                self.level_open(f'sigla_cnl: {sigla_cnl}')
                for cod_cnl_erb, ea_data in cnl_data.items():
                    self.level_open(f'cod_cnl_erb: {cod_cnl_erb}')
                    for ern, endid_data in ea_data.items():
                        self.level_open(f'ern: {ern}')
                        for g, cnl_data in endid_data.items():
                            self.level_open(f'g: {g}')
                            for ea, endid_data in cnl_data.items():
                                self.level_open(f'ea: {ea}')
                                for endid, siteid_data in endid_data.items():
                                    self.level_open(f'endid: {endid}')
                                    for siteid, cgi_data in siteid_data.items():
                                        self.level_open(f'siteid: {siteid}')
                                        for cgi, celula_data in cgi_data.items():
                                            self.level_open(f'cgi: {cgi}')
                                            for celula, ec_data in celula_data.items():
                                                self.level_open(f'celula: {celula}')
                                                for ec, erind_data in ec_data.items():
                                                    self.level_open(f'ec: {ec}')
                                                    for erind, device_data in erind_data.items():
                                                        self.level_open(f'erind: {erind}')
                                                        for device_id, device_name in device_data.items():
                                                            self.level_open(f'device_id: {device_id}={device_name}')
                                                            data.append({
                                                                'Sigla_CNL': sigla_cnl,
                                                                'Cod_cnl_ERB': cod_cnl_erb,
                                                                'ERN': ern,
                                                                'G': g,
                                                                'EA': ea,
                                                                'EndId': endid,
                                                                'SiteId': siteid,
                                                                'CGI': cgi,
                                                                'Celula': celula,
                                                                'EC': ec,
                                                                'ERIND': erind,
                                                                'idDevice': device_id,
                                                                'Device': device_name
                                                            })
                                                            self.level_close()
                                                        self.level_close()
                                                    self.level_close()
                                                self.level_close('celula')
                                            self.level_close('cgi')
                                        self.level_close('siteid')
                                    self.level_close('endid')
                                self.level_close('ea')
                            self.level_close('g')
                        self.level_close('ern')
                    self.level_close('cod_cnl_erb')
                self.level_close('sigla_cnl')
            self.level_close('reg')
            if not self.df.get(reg): self.df[reg]={}
            data=self.group_cgi_data(pd.DataFrame(data))
            self.show_done(f'{reg} Encam',data)
            self.df[reg]['cgi']=data
        self.level_close('CGI')

    def group_cgi_data(self,df:pd.DataFrame):
        """
        Agrupa dados por CGI, G, Sigla_CNL, EC, ERN
        Concatena os dispositivos em uma única coluna
        Ordena por G + Sigla_CNL
        """
        
        # Agrupar por CGI, G, Sigla_CNL, EC, ERN e concatenar os dispositivos
        grouped = df.groupby(['CGI', 'G', 'Sigla_CNL', 'EC', 'ERN']).agg({
            'Device': lambda x: ', '.join(sorted(set(x)))  # Remove duplicatas e ordena
        }).reset_index()
        grouped = grouped.rename(columns={'Device': 'Devices'})

        grouped['sort_key'] = grouped['G'].astype(str) + grouped['Sigla_CNL'].astype(str)
        grouped = grouped.sort_values('sort_key').drop('sort_key', axis=1)
        
        grouped = grouped[['CGI', 'G', 'Sigla_CNL', 'EC', 'ERN', 'Devices']]
        
        return grouped


    def excel_format_capa(self,title:str, ws: Worksheet):
        ws.delete_rows(1)
        
        larguras = {
            'A': 14, 
            'B': 14, 
            'C': 56, 
            'D': 14, 
            'E': 85,
        }
        bold = Font(color="000000", bold=True, size=11)
        normal = Font(color="000000", size=11)
        borda = Border(
            left=Side(border_style="medium", color="000000"),
            right=Side(border_style="medium", color="000000"),
            top=Side(border_style="medium", color="000000"),
            bottom=Side(border_style="medium", color="000000")
        )
        fill_Y1 = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
        fill_Y2 = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        alinhamento = Alignment(wrap_text=True, vertical='center')

        # 1. Definir largura das colunas A, B, C, D, E
        for coluna in larguras: 
            ws.column_dimensions[coluna].width = larguras[coluna]
        
        # 2. Merge das células A1:E1
        rows=[1,2,7]
        for row in rows:
            cellA=ws[f'A{row}']
            cellA.font=bold
            ws.merge_cells(f'A{row}:E{row}')
        ws.merge_cells('B3:C3')
        ws.merge_cells('B8:E8')
        ws.merge_cells('B9:E9')
        rows =range(10,ws.max_row+1)
        ini='A'
        cont=0
        for row in rows:
            cellA=ws[f'A{row}']
            val=cellA.value
            if val=='':
                ini='A'
                cont=0
            elif cont>=2:
                ini='B'
            # print(f'=============>A{row}({cont})={val}')
            if cont:
                cellA.font = bold
                for cell in ws[row]:
                    if cont==1:
                        ...
                    elif cont==2:
                        cell.font = bold
                        cell.border = borda
                        cell.fill = fill_Y1
                        cell.alignment=alinhamento
                    else:
                        cell.border = borda
                        cell.fill = fill_Y2
                        cell.alignment=alinhamento
            cells=f'{ini}{row}:E{row}'
            ws.merge_cells(cells)
            cont+=1
        
        # 3. Formatar Y1
        cells=[
            'A3', 'D3',
            'D4',
            'A5', 'B5', 'D5',
            'A6', 'B6', 'D6',
            'A8',
            'A9',
        ]
        for str_cell in cells: 
            cell=ws[str_cell]
            cell.border = borda
            cell.fill = fill_Y1
            cell.font = bold
            cell.alignment=alinhamento
            
        # 4. Formatar Y2 sem wrap
        cells=[
            'B3', 'C3', 'E3',
            'E4',
        ]
        for str_cell in cells: 
            cell=ws[str_cell]
            cell.border = borda
            cell.fill = fill_Y2
            # cell.alignment=alinhamento

        # 5. Formatar Y2 com wrap
        cells=[
            'C5', 'E5',
            'C6', 'E6',
            'B8','C8','D8','E8',
            'B9','C9','D9','E9',
        ]
        for str_cell in cells: 
            cell=ws[str_cell]
            cell.border = borda
            cell.fill = fill_Y2
            cell.alignment=alinhamento

        # 6. Ajustar altura das linhas para o conteúdo
        self.excel_fit_height_lines(ws)
        # for row in ws.iter_rows():
        #     ws.row_dimensions[row[0].row].auto_size = True
        
        return ws

    def excel_fit_height_lines(self,ws: Worksheet):
        bold=InlineFont(b=True)
        normal=InlineFont(b=False)
        for row in ws.iter_rows():
            lines = 1
            for cell in row:
                if not cell.value or not cell.alignment or not cell.alignment.wrap_text:
                    continue
                arr_str=str(cell.value).splitlines()
                change=False
                arr=[]
                sep=''
                for l in arr_str:
                    parts=l.split(sep=':',maxsplit=1)
                    if len(parts)==1:
                        arr.append(TextBlock(normal,f'{sep}{parts[0]}'))
                    else:
                        change=True
                        arr.append(TextBlock(bold,f'{sep}{parts[0]}'))
                        arr.append(TextBlock(normal,f':{parts[1]}'))
                    sep='\n'
                if change:
                    cell.value=CellRichText(*arr)
                # if cell.alignment and cell.alignment.wrap_text:
                lines = max(lines, len(arr_str))
            if lines>1:
                ws.row_dimensions[row[0].row].height = min(70,lines*16)
    
    def excel_fit_height_lines2(self,ws: Worksheet):
        for row in ws.iter_rows():
            altura_maxima = 15  # altura mínima
            for cell in row:
                if cell.value and cell.alignment and cell.alignment.wrap_text:
                    # Estimar altura baseada no conteúdo e largura da coluna
                    texto = str(cell.value)
                    largura_coluna = ws.column_dimensions[get_column_letter(cell.column)].width
                    linhas_texto = len(texto) / (largura_coluna * 0.7)  # aproximação
                    altura_estimada = max(15, linhas_texto * 11)
                    altura_maxima = max(altura_maxima, altura_estimada)
            
            ws.row_dimensions[row[0].row].height = altura_maxima
    
    def excel_create_workbook(self)->Workbook:
        wb = Workbook()
        
        wb.properties.creator = "eVoice"
        wb.properties.title = self.encam_dict['header']['dc']
        wb.properties.subject = f"Documento de Configuração: {self.encam_dict['header']['dc']}"
        # wb.properties.description = "Este documento é classificado como Uso Interno"
        # wb.properties.keywords = "uso interno, confidencial"
        
        # Remover sheet padrão
        wb.remove(wb.active)
        return wb
        
    def excel_save(self,excel_file):
        try:
            self.show(f"Salvar arquivo Excel: {excel_file}")
            self.wb.save(excel_file)
            self.show(f"- Criado com sucesso")
        except Exception as e:
            print('- '+str(e))

    def excel_insert_table(self,title:str, ws: Worksheet):
        """
        Applies formatting to the Excel worksheet and converts the data range into an Excel Table.

        Args:
            ws (str): name of table.
            ws (Worksheet): The openpyxl worksheet object.
            df (pd.DataFrame): The pandas DataFrame containing the data.
        """
        table_range = f"A1:{chr(ord('A') + ws.max_column - 1)}{ws.max_row}"
        title=f"tbl_{title.replace(' ','_')}"
        # print(f'RANGE {title}: {table_range}')
        tab = Table(displayName=title, ref=table_range)

        # Define a style for the table (similar to Excel's built-in styles)
        # This style includes header row, total row (if needed), and banded rows/columns
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style

        ws.add_table(tab)
        self.excel_worksheet_wrap_text(ws)
        self.excel_worksheet_auto_size(ws)

    def excel_worksheet_format(self,ws:Worksheet, df:pd.DataFrame):
        """
        Aplica formatação ao worksheet Excel
        """
        # Cores para formatação
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        odd_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Fonte para cabeçalho
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Fonte para dados
        data_font = Font(size=10)
        
        # Alinhamento
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Formatação do cabeçalho
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        self.excel_worksheet_wrap_text(ws)
        self.excel_worksheet_auto_size(ws)
        
    def excel_worksheet_wrap_text(self,ws:Worksheet):
        # formatr com wrap_text
        alignment = Alignment(wrap_text=True, vertical='center')
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment
    
    def excel_worksheet_auto_size(self,ws:Worksheet):
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    max_length = max(max_length,len(str(cell.value)))
                except:
                    pass
            
            ws.column_dimensions[column_letter].width = min(max_length + 2, self.excel_max_width)

    def convert_to_excel(self,sheet:str,df:pd.DataFrame)->Worksheet|None:
        """
        Função para converter dict list[dict] para Excel
        """
        
        title=self.sheets[sheet]['title']
        ws_data:Worksheet = self.wb.create_sheet(title=title)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_data.append(r)
        return ws_data

    def show_done(self,title:str,df:pd.DataFrame):
        # print(f"Colunas incluídas: {list(df.columns)}")
        # print("Preview dos primeiros 5 registros:")
        print(df.head().to_string(index=False))
        c=len(df)
        if c>5: print(f"...\nTotal de registros processados: {c}")
        print()

    def get_json_file(self,file)->dict:
        try:
            # Carregar dados JSON
            with open(file, 'r', encoding='utf-8') as file:
                cgi_data = json.load(file)
            return cgi_data
        except FileNotFoundError:
            print(f"Arquivo não encontrado: {file}")
        except json.JSONDecodeError:
            print("Erro ao decodificar JSON. Verifique se o arquivo está no formato correto.")
        except Exception as e:
            print(f"Erro inesperado: {str(e)}")
        return {}

    def show(self,text):
        if not self.verbose: return
        print(text)
    def level_open(self,text):
        self.evel_item(text,'- ')
        self.level+=1
    def evel_item(self,text,c=''):
        if self.verbose<2 or self.level<0: return
        if self.verbose==2:
            print(f'{"":<{self.level*2}}{c}{text}')
        else:
            print(f'{"":<{self.level*2}}{c}[{self.level}]{text}')
    def level_close(self,text=None):
        self.level-=1
        if self.verbose>2:
            self.evel_item(f'close {text}','= ')
        
if __name__ == "__main__":
    DC_SUP()